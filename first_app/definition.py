import pyvisa
import sys
import os
import math
import openpyxl
# import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import serial
import re
# import config
import datetime
from decimal import * # should aviod wildcard import mentioned in PEP08

getcontext().prec = 10 # set 10 decimal values precision

# Decimal module import from decimal make sure float numbers addition yeilds correct value
import sounddevice as sd
import numpy as np
import matplotlib.pyplot as plt
from scipy.fftpack import fft
import subprocess
from os import system
from subprocess import Popen, PIPE
from first_app.models import FEP, RES, DEMOD, TXSIG, ACP, CSE, CSEOUT, CSHOUT, RXSIG

class SoundCard(object):
    def __init__(self):
        self.gain = 10
        self.duration = 0.1 # sample record duration
        self.sample_rate = 44100
        self.device = ('Microphone (Sound Blaster Play!, MME',
                       'Speakers (Sound Blaster Play! 3, MME')
        sd.query_devices(self.device[0])
        sd.default.device = self.device
        sd.default.channels = 1
        sd.default.samplerate = self.sample_rate

    def get_sample(self, average=5):
        x4 = np.full(3528, 0.0)
        for i in range(0, average):
            x1 = sd.rec(int(self.duration*self.sample_rate), dtype='float32', blocking=True) # record 0.1 seconds input signal from microphone

            x2 = self.gain*x1[int(self.duration*self.sample_rate/1000):int(self.duration*self.sample_rate)] # take effective samples

            x3 = x2[0:int(80*(self.sample_rate/1000))] # take first 80ms duration samples

            x4 = x3.flatten()# transfer 2D array to 1D, ready for fft operation, fft function can only take 1D array as input


        x4 = x4 / average
        # print(x4)


        t = 80*np.linspace(0, 1, np.size(x4))# generate 80ms time axis VS samples for first 80ms

        X = fft(x4) # fft transform
        n = np.size(t)
        self.fr = (self.sample_rate/2)*np.linspace(0, 1, int(n/2)) # generate frequecny axis array

        X_m = (2/n)*abs(X[0:np.size(self.fr)])

        # print(np.size(X_m))# M=2*np.size(X_m)=3528
        # print(X_m)

        return {'Time_level':x4, 'Time':t, 'Freq_level':X_m, 'Freq':self.fr}

    def get_SINAD(self, dict): # accept returned dictionary from get_sample()
        X_mmax = np.max(dict['Freq_level']) # find maximum value in numpy array
        # print(X_mmax)
        fr_index = np.where(dict['Freq_level'] == X_mmax) # locate index in numpy array
        fr_max = dict['Freq'][fr_index] # find corresponding frequency of maximum value

        # print(fr_max)
        # print(fr_index)
        # print(fr_max)

        SND = np.sum(dict['Freq_level'][24:240]) # from 300Hz to 3kHz get SND
        # print(SND)
        mask = self.mask_gen(x=fr_index[0], y=SND, z=self.fr)
        ND = np.sum(dict['Freq_level'][24:240] * mask[24:240]) # from 300Hz to 3kHz get ND
        # print(ND)
        # 1kHz at index 80
        #freq bin width = fs/M = 44100/3528 = 12.5Hz
        # 300/12.5 = 24 meaning 300Hz at index 24
        # 3000/12.5 = 240 meaning 3kHz at index 240
        SINAD = 20 * np.log10(SND/ND) # calculate SINAD
        return SINAD

    def mask_gen(self, x, y, z):
        mask = np.full(np.size(z), 1)
        mask[x]=0
        if y>=9.0:
            for i in range(1,30):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 30")
        elif y>=8.0:
            for i in range(1,40):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 40")
        elif y>=7.0:
            for i in range(1,55):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 55")
        elif y>=6.0:
            for i in range(1,70):
                mask[x-i]=0
                mask[x+i]=0
                print("filter bin set to 70")
        elif y>=5.0:
            for i in range(1,85):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 85")
        elif y>=4.0:
            for i in range(1,100):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 100")
        elif y>=3.0:
            for i in range(1,110):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 110")
        elif y>=2.0:
            for i in range(1,120):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 120")
        else:
            for i in range(1,130):
                mask[x-i]=0
                mask[x+i]=0
            print("filter bin set to 130")

        return mask


    def live_plots(self):
        dict = self.get_sample()
        fig, (ax0, ax1) = plt.subplots(nrows=2)
        ax0.set_title('Sinusoidal Signal')
        ax0.set_xlabel('Time(ms)')
        ax0.set_ylabel('Amplitude')
        ax0.set_ylim(-2, 2)
        line0, = ax0.plot(dict['Time'], dict['Time_level']) # time VS level plot

        ax1.set_title('Magnitude Spectrum')
        ax1.set_xlabel('Frequency(Hz)')
        ax1.set_ylabel('Magnitude')
        ax1.set_ylim(0, 1.5)
        ax3 = ax1.text(0.3, 0.9, 'SINAD', transform=ax1.transAxes)
        line1, = ax1.plot(dict['Freq'], dict['Freq_level']) # frequency VS level plot

        fig.subplots_adjust(hspace=0.6)

        while True:
            dict = self.get_sample()
            line0.set_ydata(dict['Time_level']) # update data in plots only in loop instead of update entire plots
            line1.set_ydata(dict['Freq_level'])
            ax3.set_text('SINAD = %.2f dB' % self.get_SINAD(dict=dict)) # update SINAD value

            plt.pause(0.01)# this command will call plt.show()

class SigGen(object):

    def __init__(self, address):
        self.rm = pyvisa.ResourceManager()
        self.SG = self.rm.open_resource(address)

    def Set_Timeout(self, ms):
        self.SG.timeout = ms # useful on CMS for Rx test, pyvisa parameter

    def Lev_AF(self):
        return self.Level_AF # useful for Max_deviation test

    def Lev_RF(self):
        return self.SG.query(':POW?') # useful for Rx test

    def Freq_RF(self):
        return self.SG.query(':FREQ?') # useful for Rx test

    def write(self, str):
        self.SG.write(str)

    def query(self, str):
        return self.SG.query(str)

    def close(self):
        self.SG.close()

    def Tx_Setup(self):
        txsig_list = TXSIG.objects.all()
        self.Level_AF = txsig_list[1].content

        self.SG.write(f"*RST") # write all paramaters to SigGen
        self.SG.write("SYST:DISP:UPD ON")
        #self.SG.write(f":FM:INT:FREQ {Frequency_AF}kHz")
        self.SG.write(f"LFO:FREQ {txsig_list[0].content}kHz")
        # self.SG.write(f":OUTP2:VOLT {self.Level_AF}mV")
        self.SG.write(f"LFO:VOLT {self.Level_AF}mV")
        # self.SG.write(f":OUTP2 {AF_output_on}")
        self.SG.write(f"LFO {txsig_list[2].content}")

    def Unwanted_Signal(self, freq):
        rxsig_list = RXSIG.objects.all()

        self.ULevel_RF = rxsig_list[7].content

        self.SG.write(f"*RST") # write all paramaters to SigGen
        self.SG.write("SYST:DISP:UPD ON")
        self.SG.write(f"FREQ {freq}MHz")
        self.SG.write(f":UNIT:POW dBuV")
        self.SG.write(f":POW {self.ULevel_RF}dBuV")
        self.SG.write(f":FM:INT:FREQ {rxsig_list[8].content}Hz")
        self.SG.write(f"FM {rxsig_list[9].content}kHz")
        self.SG.write(f":FM:STAT {rxsig_list[10].content}")
        self.SG.write(f":OUTP1 {rxsig_list[11].content}")

    def Wanted_Signal(self, freq):
        rxsig_list = RXSIG.objects.all()

        self.WLevel_RF = rxsig_list[1].content

        self.SG.write(f"*RST") # write all paramaters to SigGen
        self.SG.write("SYST:DISP:UPD ON")
        self.SG.write(f"FREQ {freq}MHz")
        self.SG.write(f":UNIT:POW dBuV")
        self.SG.write(f":POW {self.WLevel_RF}dBuV")
        self.SG.write(f"LFO:FREQ {rxsig_list[2].content}kHz")
        self.SG.write(f"FM {rxsig_list[3].content}kHz")
        self.SG.write(f"FM:STAT {rxsig_list[4].content}")
        self.SG.write(f"OUTP {rxsig_list[5].content}")


class SpecAn(object):
    def __init__(self, address):
        self.rm = pyvisa.ResourceManager()
        self.SP = self.rm.open_resource(address) # Spec An

    def write(self, str):
        self.SP.write(str)

    def query(self, str):
        return self.SP.query(str)

    def close(self):
        self.SP.close()

    def screenshot(self, file_name, folder):
        self.SP.write("HCOP:DEV:LANG PNG") # set file type to .png
        self.SP.write("HCOP:CMAP:DEF4")
        self.SP.write(f"MMEM:NAME \'c:\\temp\\Dev_Screenshot.png\'")
        self.SP.write("HCOP:IMM") # perform copy and save .png file on SpecAn
        self.SP.query("*OPC?")

        file_data = self.SP.query_binary_values(f"MMEM:DATA? \'c:\\temp\\Dev_Screenshot.png\'", datatype='s',)[0] # query binary data and save
        # new_file = open(f"c:\\Temp\\{file_name}.png", "wb")
        new_file = open(f"c:\\Users\\afan\\Documents\\titan\\static\\result_images\\{folder}\\{file_name}.png", "wb")# open a new file as "binary/write" on PC
        new_file.write(file_data) # copy data to the file on PC
        new_file.close()
        print(f"Screenshot saved to PC c:\\Users\\afan\\Documents\\titan\\static\\result_images\\{folder}\\{file_name}.png\n")

    def FEP_Setup(self, freq):
        fep_list = FEP.objects.all()
        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write(f"FREQ:CENT {freq}MHz")
        self.SP.write(f"FREQ:SPAN {fep_list[1].content}Hz")
        self.SP.write(f"BAND {fep_list[2].content}Hz")
        self.SP.write(f"BAND:VID {fep_list[3].content}Hz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {fep_list[6].content}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {fep_list[4].content}")
        self.SP.write(f"INP:ATT {fep_list[5].content}")
        self.SP.write(f"{fep_list[7].content}")
        self.SP.write(f"CORR:TRAN:SEL '{fep_list[8].content}'")
        self.SP.write(f"CORR:TRAN {fep_list[9].content} ")
        self.SP.write(f"CORR:TRAN:SEL '{fep_list[10].content}'")
        self.SP.write(f"CORR:TRAN {fep_list[11].content}")
        self.SP.write(f"CORR:TRAN:SEL '{fep_list[12].content}'")
        self.SP.write(f"CORR:TRAN {fep_list[13].content}")
        self.SP.write(f"CALC:LIM:NAME '{fep_list[14].content}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {fep_list[15].content}")
        self.SP.write(f"CALC:LIM:NAME '{fep_list[16].content}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {fep_list[17].content}")
        self.SP.write(f"SWE:POIN {fep_list[18].content}")

    def get_FEP_result(self, freq, folder):
        self.SP.write("CALC:MARK1:MAX")
        Frequency = self.SP.query("CALC:MARK1:X?")
        Level = float(self.SP.query("CALC:MARK1:Y?"))
        Frequency_error = float(Frequency) - float(self.SP.query("FREQ:CENT?"))
        indication = (self.SP.query("*OPC?")).replace("1","Completed.")
        self.screenshot('FEP_'+str(freq)+'_MHz', folder)
        return {'F':Frequency_error, 'F_limit':1500, 'F_margin':1500 - abs(Frequency_error),
                'P':Level, 'P_limit':36.98, 'P_margin':abs(Level - 36.98),
                'Screenshot': 'FEP_'+str(freq)+'_MHz'
                }


    def get_CSE_result(self, freq, sub_range):
        self.SP.write("CALC:MARK1:MAX")
        self.SP.write("CALC:MARK2:MAX")
        self.SP.write("CALC:MARK2:MAX:NEXT")
        Frequency1 = float(self.SP.query("CALC:MARK1:X?"))/1e6
        Level1 = float(self.SP.query("CALC:MARK1:Y?"))

        Frequency2 = float(self.SP.query("CALC:MARK2:X?"))/1e6
        Level2 = float(self.SP.query("CALC:MARK2:Y?"))

        if sub_range <= 3:
            cse_list = CSEOUT.objects.get_or_create(SubRange=sub_range,
                                                    CSE1_Frequency_MHz=round(Frequency1,5),
                                                    CSE1_Level_dBm=round(Level1,5),
                                                    CSE2_Frequency_MHz=round(Frequency2,5),
                                                    CSE2_Level_dBm=round(Level2,5),
                                                    limit_dBm= -30,
                                                    Screenshot_file='CSE0'+str(sub_range)+'_'+str(freq)+'_MHz.png'
                                                    )[0]
            indication = (self.SP.query("*OPC?")).replace("1","Completed.")
            print(f"CSE Test {indication}")
        else:
            cse_list = CSHOUT.objects.get_or_create(SubRange=sub_range,
                                                    CSE1_Frequency_MHz=round(Frequency1,5),
                                                    CSE1_Level_dBm=round(Level1,5),
                                                    CSE2_Frequency_MHz=round(Frequency2,5),
                                                    CSE2_Level_dBm=round(Level2,5),
                                                    limit_dBm= -30,
                                                    Screenshot_file='CSE0'+str(sub_range)+'_'+str(freq)+'_MHz.png'
                                                    )[0]
            indication = (self.SP.query("*OPC?")).replace("1","Completed.")
            print(f"CSE Test {indication}")

        return cse_list


    def CSE_Setup(self, sub_range, limit_line, cutoff, filter):

        if limit_line == 1:
            CSE.objects.filter(id=16).update(content='OFF')
            CSE.objects.filter(id=18).update(content='ON')
            CSE.objects.filter(id=21).update(content='OFF')

        if sub_range == 1:# choose sub_range
            CSE.objects.filter(id=1).update(content=0.009)
            CSE.objects.filter(id=2).update(content=0.15)
            CSE.objects.filter(id=3).update(content=1)
            CSE.objects.filter(id=4).update(content=3)
            CSE.objects.filter(id=5).update(content=15)
            CSE.objects.filter(id=6).update(content=25)
            CSE.objects.filter(id=7).update(content=3.5)
            CSE.objects.filter(id=10).update(content='ON')
            CSE.objects.filter(id=12).update(content='ON')
            CSE.objects.filter(id=14).update(content='OFF')
            CSE.objects.filter(id=19).update(content=5001)


        elif sub_range == 2:
            CSE.objects.filter(id=1).update(content=0.15)
            CSE.objects.filter(id=2).update(content=30)
            CSE.objects.filter(id=3).update(content=10)
            CSE.objects.filter(id=4).update(content=30)
            CSE.objects.filter(id=5).update(content=15)
            CSE.objects.filter(id=6).update(content=25)
            CSE.objects.filter(id=7).update(content=3.5)
            CSE.objects.filter(id=10).update(content='ON')
            CSE.objects.filter(id=12).update(content='ON')
            CSE.objects.filter(id=14).update(content='OFF')
            CSE.objects.filter(id=19).update(content=5001)


        elif sub_range == 3:
            CSE.objects.filter(id=1).update(content=30)
            CSE.objects.filter(id=2).update(content=cutoff)
            CSE.objects.filter(id=3).update(content=100)
            CSE.objects.filter(id=4).update(content=300)
            CSE.objects.filter(id=5).update(content=50)
            CSE.objects.filter(id=6).update(content=25)
            CSE.objects.filter(id=7).update(content=33.8)
            CSE.objects.filter(id=10).update(content='ON')
            CSE.objects.filter(id=12).update(content='OFF')
            CSE.objects.filter(id=14).update(content='OFF')
            CSE.objects.filter(id=19).update(content=10001)


        elif sub_range == 4:
            CSE.objects.filter(id=1).update(content=cutoff)
            CSE.objects.filter(id=2).update(content=1000)
            CSE.objects.filter(id=3).update(content=100)
            CSE.objects.filter(id=4).update(content=300)
            CSE.objects.filter(id=5).update(content=0)
            CSE.objects.filter(id=6).update(content=10)
            CSE.objects.filter(id=7).update(content=3.5)
            CSE.objects.filter(id=10).update(content='ON')
            CSE.objects.filter(id=12).update(content='ON')
            CSE.objects.filter(id=13).update(content=filter)
            CSE.objects.filter(id=14).update(content='ON')
            CSE.objects.filter(id=19).update(content=20001)


        elif sub_range == 5:
            CSE.objects.filter(id=1).update(content=1000)
            CSE.objects.filter(id=2).update(content=4000)
            CSE.objects.filter(id=3).update(content=1000)
            CSE.objects.filter(id=4).update(content=3000)
            CSE.objects.filter(id=5).update(content=-10)
            CSE.objects.filter(id=6).update(content=5)
            CSE.objects.filter(id=7).update(content=3.5)
            CSE.objects.filter(id=10).update(content='ON')
            CSE.objects.filter(id=12).update(content='ON')
            CSE.objects.filter(id=13).update(content=filter)
            CSE.objects.filter(id=14).update(content='ON')
            CSE.objects.filter(id=19).update(content=30001)


        else:
            pass

        cse_list = CSE.objects.all()
        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write(f"FREQ:STAR {cse_list[0].content}MHz")
        self.SP.write(f"FREQ:STOP {cse_list[1].content}MHz")
        self.SP.write(f"BAND {cse_list[2].content}kHz")
        self.SP.write(f"BAND:VID {cse_list[3].content}kHz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {cse_list[6].content}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {cse_list[4].content}")
        self.SP.write(f"INP:ATT {cse_list[5].content}")
        self.SP.write(f"{cse_list[6].content}")
        self.SP.write(f"{cse_list[7].content}")
        self.SP.write(f"CORR:TRAN:SEL '{cse_list[8].content}'")
        self.SP.write(f"CORR:TRAN {cse_list[9].content} ")
        self.SP.write(f"CORR:TRAN:SEL '{cse_list[10].content}'")
        self.SP.write(f"CORR:TRAN {cse_list[11].content}")
        self.SP.write(f"CORR:TRAN:SEL '{cse_list[12].content}'")
        self.SP.write(f"CORR:TRAN {cse_list[13].content}")
        self.SP.write(f"CALC:LIM:NAME '{cse_list[14].content}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {cse_list[15].content}")
        self.SP.write(f"CALC:LIM:NAME '{cse_list[16].content}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {cse_list[17].content}")
        self.SP.write(f"SWE:POIN {cse_list[18].content}")
        self.SP.write(f"CALC:LIM:NAME '{cse_list[19].content}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {cse_list[20].content}")
        self.SP.write(f"DISP:TRAC:MODE MAXH")




    def ACP_Setup(self, freq):
        acp_list = ACP.objects.all()

        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write("CALC:MARK:FUNC:POW:SEL ACP")
        self.SP.write(f"FREQ:CENT {freq}MHz")
        self.SP.write(f"FREQ:SPAN {acp_list[1].content}kHz")
        self.SP.write(f"BAND {acp_list[2].content}Hz")
        self.SP.write(f"BAND:VID {acp_list[3].content}Hz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {acp_list[6].content}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {acp_list[4].content}")
        self.SP.write(f"INP:ATT {acp_list[5].content}")
        self.SP.write(f"{acp_list[7].content}")
        self.SP.write(f"POW:ACH:BWID:CHAN1 {acp_list[8].content}kHz")
        self.SP.write(f"POW:ACH:BWID:ACH {acp_list[9].content}kHz")
        self.SP.write(f"POW:ACH:BWID:ALT1 {acp_list[10].content}kHz")
        self.SP.write(f"POW:ACH:ACP {acp_list[11].content}")
        self.SP.write(f"POW:ACH:SPAC {acp_list[12].content}kHz")
        self.SP.write(f"POW:ACH:SPAC:ALT1 {acp_list[13].content}kHz")
        self.SP.write(f"POW:ACH:MODE {acp_list[14].content}")
        self.SP.write(f"SWE:COUN {acp_list[15].content}")
        self.SP.write(f"CALC:MARK:FUNC:POW:MODE WRIT")
        self.SP.write(f"DISP:TRAC:MODE AVER")

    def DeMod_Setup(self, freq):
        demod_list = DEMOD.objects.all()

        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write("ADEM ON")
        self.SP.write(f"FREQ:CENT {freq}MHz")
        self.SP.write(f"DISP:TRAC:Y:PDIV {demod_list[1].content}kHz")
        self.SP.write(f"BAND:DEM {demod_list[2].content}kHz")
        self.SP.write(f"ADEM:AF:COUP {demod_list[3].content}")
        #Set AF coupling to AC, the frequency offset is automatically corrected.
        # i.e. the trace is always symmetric with respect to the zero line
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {demod_list[6].content}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {demod_list[4].content}")
        self.SP.write(f"INP:ATT {demod_list[5].content}")
        self.SP.write(f"{demod_list[7].content}")
        self.SP.write(f"ADEM:MTIM {demod_list[8].content}ms")
        self.SP.write(f"INIT:CONT {demod_list[9].content}")


class Radio(object):
    def __init__(self, com, baudrate):
        self.ESC_CHAR = 0x7D # Escape char
        self.STX_CHAR =  0x7E # Start of packet
        self.ETX_CHAR = 0x7F # End of packet
        self.CHECKSUM_XOR_MASK = 0xFF

        self.payload0 = bytearray(b'\xB4\x88\x2a\x80')# set carrie freq 136MHz basd on 12.5kHz calculation
        self.payload1 = bytearray(b'\xB4\x93\x03')# set 25W power
        self.payload2 = bytearray(b'\xA6\x01') # PTT ON
        self.payload3 = bytearray(b'\xB4\x91\x03\xE8')# 1kHz audio on
        self.payload4 = bytearray(b'\xB4\x91\x0B\xB8')# 3kHz audio on
        self.payload5 = bytearray(b'\xB4\x91\x00\x00')# audio off
        self.payload6 = bytearray(b'\xB4\x82\x10')# selcall tone on
        self.payload7 = bytearray(b'\xB4\x82\x0f')# selcall tone off
        self.payload8 = bytearray(b'\xA6\x00') # PTT off

        self.port = com
        self.baudrate = baudrate
        self.timeout = None
        self.ser = serial.Serial(port=self.port, baudrate=self.baudrate, timeout=self.timeout)  # open serial port

    def Packet_Gen(self, payload): # return GME2 protocol packet
        # Header
        tx_array = bytearray((self.STX_CHAR, 2, len(payload)))

        # Escape the payload, append it to the output buffer, make the checksum
        chksum = 0
        for i, b in enumerate(payload):
            # The first word of the payload is always the command ID.
            # The MSB of the command ID is always 1. The radio just
            # clears it in the ACK, without doing anything else with it.
            if i == 0:
                bb = (0x80 | b) & 0xFF
            else:
                bb = b & 0xFF
            chksum ^= bb
            if bb in (self.ESC_CHAR, self.STX_CHAR, self.ETX_CHAR):
                tx_array.append(self.ESC_CHAR)
                tx_array.append(bb ^ 0x20)
            else:
                tx_array.append(bb)

        # Trailer
        tx_array.append(chksum ^ self.CHECKSUM_XOR_MASK)
        tx_array.append(self.ETX_CHAR)

        return tx_array

    def Set_Freq(self, freq):
        a = ((hex(int((Decimal(freq)*Decimal(1e6))/(Decimal(12.5*1e3))))[2]+hex(int((Decimal(freq)*Decimal(1e6))/Decimal((12.5*1e3))))[3]))# calculate first HEX of input frequency
        b = ((hex(int((Decimal(freq)*Decimal(1e6))/(Decimal(12.5*1e3))))[4]+hex(int((Decimal(freq)*Decimal(1e6))/Decimal((12.5*1e3))))[5]))# calculate second HEX of input frequency
        self.payload0[2] = int(a,16) # assign first DEC (transferred from HEX) to the third number of paylaod0
        self.payload0[3] = int(b,16) # assign second DEC (transferred from HEX) to the fourth number of paylaod0
        self.ser.write(self.Packet_Gen(self.payload0))
        print(f"radio frequency has been set to {freq} MHz.")



    def Set_Pow(self, pow):
        if pow == "low":
            self.payload1[2] = 0
        elif pow == "high":
            self.payload1[2] = 3
        else:
            print("Set_Pow() only accept 'low' or 'high'")
        print(f"radio power has been set to {pow}.")

        self.ser.write(self.Packet_Gen(self.payload1))

    def Radio_On(self):
        self.ser.write(self.Packet_Gen(self.payload2))
        print("Radio's on")

    def Radio_Off(self):
        self.ser.write(self.Packet_Gen(self.payload8))
        print("Radio's off")

    def Radio_close(self):
        self.ser.close()
        print("Serial session is closed.")

class Excel(object):

    def __init__(self, file_name):
        self.file = load_workbook(filename = file_name) # load Test_Result.xlsx

    def get_sheet(self, sheet_name):
        self.sheet = self.file[sheet_name]

    def write(self, row, column, value):
        self.sheet.cell(row = row, column = column, value = value)

    def clear(self, start_cell, end_cell):
        for row in self.sheet[start_cell +':'+ end_cell]:# clear certain block of cells in selected sheet
            for cell in row:
                cell.value = None
    def save(self, file_name):
        self.file.save(file_name)



def Tx_set_standard_test_condition():
    Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0 #get the initial deviation value
    Level_AF = float(SMB1.Lev_AF())# initial Level_AF
# above code is to find Audio output level
# satisfying standard condtion (around 1.5kHz deviation)
    if Dev_Reading < 1.5:
        while Dev_Reading < 1.47:
            print(f"current deviation:{Dev_Reading}kHz")
            Level_AF = Level_AF+1
            SMB1.write(f"LFO:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    else:
        while Dev_Reading > 1.53:
            print(f"current deviation:{Dev_Reading}kHz")
            Level_AF = Level_AF-1
            SMB1.write(f"LFO:VOLT  {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0

    FSV.write(f"INIT:CONT ON")
    FSV.query('*OPC?')
    print(f"Audio level has been set to {Level_AF} mV")
    return Level_AF

def CSE_operation(freq, sub_range, limit_line, cutoff, filter='NA'):
    FSV.CSE_Setup(sub_range=sub_range, limit_line=limit_line, cutoff=cutoff, filter=filter)
    FSV.query('*OPC?')
    EUT.Set_Freq(freq=freq+0.0125)
    EUT.Set_Pow("high")
    EUT.Radio_On()
    time.sleep(5)
    FSV.write("DISP:TRAC:MODE VIEW")
    EUT.Radio_Off()
    FSV.get_CSE_result(freq=freq, sub_range=sub_range)
    FSV.screenshot(file_name='CSE0'+str(sub_range)+'_'+str(freq)+'_MHz', folder='cs')

def ACS_operation(freq, delta):
    EUT.Set_Freq(freq=freq+0.0125)
    SMB1.Wanted_Signal(freq=freq)
    SMB1.query('*OPC?')
    SMB2.Unwanted_Signal(freq=freq+delta)
    SMB2.query('*OPC?')

    SINAD = SC.get_SINAD(dict=SC.get_sample())
    # get initial SINAD from SC
    print(SINAD)
    Level_RF = SMB2.Lev_RF()
    for i in range(0,100):
        if SINAD > 14.0:
            Level_RF = float(Level_RF) + 1
            SMB2.write(f":POW {Level_RF}dBuV")
            SMB2.query('*OPC?')
            SINAD = SC.get_SINAD(dict=SC.get_sample())
            print(SINAD)
        else:
            break
    ACS = Level_RF - float(SMB1.Lev_RF())
    return ACS





equip_list = RES.objects.all()

try:
    FSV = SpecAn(equip_list[2].resadd)
except BaseException:
    print("FSV is not on.")
    pass

try:
    SMB1 = SigGen(equip_list[0].resadd)
except BaseException:
    print("SMB1 is not on.")
    pass

try:
    SMB2 = SigGen(equip_list[1].resadd)
except BaseException:
    print("SMB2 is not on.")
    pass

try:
    EUT = Radio('com10', baudrate=115200)
except BaseException:
    print("Specified com port does not exsit.")
    pass

try:
    SC = SoundCard()
except BaseException:
    print("Specified Soundcard does not exist.")
    pass

# def equipment_assignment(FSVIP=equip_list[0].resadd, SMB1IP=equip_list[1].resadd,
#                         SMB2IP=equip_list[2].resadd, com='com10', baudrate=9600
#                         ):
#
#     try:
#         FSV = SpecAn(FSVIP)
#     except BaseException:
#         print("FSV is not on.")
#         pass
#
#     try:
#         SMB1 = SigGen(SMB1IP)
#     except BaseException:
#         print("SMB1 is not on.")
#         pass
#
#     try:
#         SMB2 = SigGen(SMB2IP)
#     except BaseException:
#         print("SMB2 is not on.")
#         pass
#
#     try:
#         EUT = Radio(com, baudrate)
#     except BaseException:
#         print("Specified com port does not exsit.")
#         pass
#
#     list = [FSV, SMB1, SMB2, EUT]
#
#     return list
#
# equipment_assignment()
