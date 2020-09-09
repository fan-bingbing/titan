import visa
from openpyxl import load_workbook
import time
import re
import modules
import datetime












































# payload0 = bytearray(b'\xB4\x88\x2a\x80')# set carrie freq 136MHz basd on 12.5kHz calculation
payload0 = bytearray(b'\xB4\x88\x7D\x00')# set carrie freq 400MHz basd on 12.5kHz calculation
payload1 = bytearray(b'\xB4\x93\x03')# set 25W power
payload2 = bytearray(b'\xA6\x01') # PTT ON
payload3 = bytearray(b'\xB4\x91\x03\xE8')# 1kHz audio on
payload4 = bytearray(b'\xB4\x91\x0B\xB8')# 3kHz audio on
payload5 = bytearray(b'\xB4\x91\x00\x00')# audio off
payload6 = bytearray(b'\xB4\x82\x10')# selcall tone on
payload7 = bytearray(b'\xB4\x82\x0f')# selcall tone off
payload8 = bytearray(b'\xA6\x00') # PTT off

# naming rules for excel:
# 1. for operating variables related to Test_Setup.xlsx: use SFile_write, SSheet
# 2. for operating variables related to Test_Result.xlsx: use RFile_write, RSheet

Start_F = int(input("Input start frequency in Mhz or press CTRL+C to quit > "))
Stop_F = int(input("Input stop frequency in Mhz or press CTRL+C to quit > "))
steps = int((Stop_F-Start_F)*1000/12.5)

rm = visa.ResourceManager()
FSV = rm.open_resource('TCPIP0::10.0.22.30::hislip0::INSTR') # Spec An

# below codes are for setting test frequency in Test_Setup.xlsx according to user's input
FSV_file_write = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file
SSheet = FSV_file_write["ACP"] # load existing SSheet named "ACP"
SSheet.cell(row = 1, column = 2, value = Start_F) # write test frequency in this SSheet
FSV_file_write.save("Test_Setup.xlsx") # save existing .xlsx file
# above codes are for setting test frequency in Test_Setup.xlsx according to user's input

FSV.clear()

FSV_file_write = load_workbook(filename = "Test_Setup.xlsx") # create a workbook from existing .xlsx file
SSheet = FSV_file_write["ACP"] # load setup SSheet in .xlsx to SSheet



RFile_write = load_workbook(filename = "CM60_Result.xlsx") # load Test_Result.xlsx
RSheet = RFile_write["Radio1"] # load "ACP" sheet in .xlsx




# following code is to initialize FSV to complete the test
Centre_frequency = SSheet["B1"].value # get all parameters from Test_Setup.xlsx
Span_frequency = SSheet["B2"].value #
RBW = SSheet["B3"].value #
VBW = SSheet["B4"].value #
RF_level = SSheet["B5"].value #
Attenuation = SSheet["B6"].value #
RefLev_offset = SSheet["B7"].value#
Trace_RMS = SSheet["B8"].value #

Tx_CHBW = SSheet["B10"].value
AJ_CHBW = SSheet["B11"].value
AT_CHBW = SSheet["B12"].value
AJ_CHNUM = SSheet["B13"].value
AJ_SPACE = SSheet["B14"].value
AT_SPACE = SSheet["B15"].value
Power_Mode = SSheet["B16"].value
Ave_number = SSheet["B17"].value

FSV.write(f"*RST")
FSV.write("SYST:DISP:UPD ON")
FSV.write("CALC:MARK:FUNC:POW:SEL ACP")

FSV.write(f"FREQ:CENT {Centre_frequency}MHz") # set all parameters
FSV.write(f"FREQ:SPAN {Span_frequency}kHz")
FSV.write(f"BAND {RBW}Hz")
FSV.write(f"BAND:VID {VBW}Hz")
FSV.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
FSV.write(f"INP:ATT {Attenuation}")
FSV.write(f"DISP:TRAC:Y:RLEV {RF_level}")
FSV.write(f"{Trace_RMS}")

FSV.write(f"POW:ACH:BWID:CHAN1 {Tx_CHBW}kHz")
FSV.write(f"POW:ACH:BWID:ACH {AJ_CHBW}kHz")
FSV.write(f"POW:ACH:BWID:ALT1 {AT_CHBW}kHz")
FSV.write(f"POW:ACH:ACP {AJ_CHNUM}")
FSV.write(f"POW:ACH:SPAC {AJ_SPACE}kHz")
FSV.write(f"POW:ACH:SPAC:ALT1 {AT_SPACE}kHz")
FSV.write(f"POW:ACH:MODE {Power_Mode}")
FSV.write(f"SWE:COUN {Ave_number}")
FSV.write(f"CALC:MARK:FUNC:POW:MODE WRIT")
#FSV.write(f"DISP:TRAC:MODE AVER")
time.sleep(3)
#FSV.write(f"DISP:TRAC:MODE VIEW")
FSV.query("*OPC?")

CM60_Control.Radio_Con('com10', 9600, None, payload1)# set radio to 25W


for i in range(0,steps):

    CM60_Control.Radio_Con('com10', 9600, None, payload0)# set frequency

    CM60_Control.Radio_Con('com10', 9600, None, payload2)# PTT on
    CM60_Control.Radio_Con('com10', 9600, None, payload3)# set audio
    time.sleep(1)
    FSV.write(f"DISP:TRAC:MODE AVER")
    time.sleep(3)

    ACP = FSV.query("CALC:MARK:FUNC:POW:RES? ACP")
    LIST = re.findall(r'-?\d+\.\d+', ACP)
    Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())


    print(LIST)


    RSheet.cell(row = i+2, column = 1, value = Centre_frequency)
    RSheet.cell(row = i+2, column = 2, value = float(LIST[0]))
    RSheet.cell(row = i+2, column = 3, value = float(LIST[1]))
    RSheet.cell(row = i+2, column = 4, value = float(LIST[2]))
    RSheet.cell(row = i+2, column = 5, value = Timestamp)
    RFile_write.save("CM60_Result.xlsx") # save existing .xlsx file
    FSV.query("*OPC?")
    CM60_Control.Radio_Con('com10', 9600, None, payload8) # PTT off

    if payload0[3]<=254: # payload stores decimal value ranging from 0 to 255
        payload0[3] = payload0[3] + 1 # carrier frequency add 12.5kHz per iteration
    else:
        payload0[3] = 0
        payload0[2] = payload0[2] + 1

    print("Current Channel HEX:")
    print(hex(payload0[2]),hex(payload0[3]))


    # str1 = "".join((str(hex(payload0[2])),str(hex(payload0[3]))))
    # print(int(str1,16))
    # print(f"Current Channel frequency:{int(str1,16)}")

    Centre_frequency = float(Centre_frequency) + 12.5/1000
    FSV.write(f"FREQ:CENT {Centre_frequency}MHz")
    FSV.write(f"DISP:TRAC:MODE WRIT")

#print(FSV.query("*OPC?")).replace("1","ACP test Completed") # replace return character "1" to "completed"

FSV.close()
